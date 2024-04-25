from pydantic import BaseModel
import mysql.connector
import random
import logging
import string
from pydantic import BaseModel
#from datetime import date, time
#import datetime
from datetime import datetime, date, time, timedelta
import pandas as pd
import io
import tempfile

from typing import List
import openpyxl
from io import BytesIO
from fastapi import FastAPI, File, UploadFile, HTTPException, Depends

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware

from tempfile import NamedTemporaryFile

###################### ENCRYPTION
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend
import base64

def encrypt_with_aes_ecb(plaintext):
    key = "segproject"        
    # Convert the string key to bytes and pad it to make it 32 bytes long
    key_bytes = key.ljust(32, "\x00").encode()

    # Create a Cipher object with AES in ECB mode
    cipher = Cipher(algorithms.AES(key_bytes), modes.ECB(), backend=default_backend())

    # Pad the plaintext to be a multiple of the block size (16 bytes for AES)
    padding_length = 16 - (len(plaintext) % 16)
    padded_plaintext = plaintext + bytes([padding_length]) * padding_length

    # Encrypt the plaintext
    encryptor = cipher.encryptor()
    ciphertext = encryptor.update(padded_plaintext) + encryptor.finalize()

    # Encode the ciphertext as URL-safe base64
    encrypted_text = base64.urlsafe_b64encode(ciphertext)

    return encrypted_text.decode()

app = FastAPI()
#Allows all origins, methods, and headers. Use with caution and only in development.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allows all origins
    allow_credentials=True,
    allow_methods=["*"],  # Allows all methods
    allow_headers=["*"],  # Allows all headers
)

# Define database connection settings
MYSQL_CONFIG = {
    "host": "127.0.0.1",
    "port": 3306,
    "user": "root",
    "password": "",
    "database": "rbms2"     #CHANGE THIS
}

# Dependency to establish database connection
def get_database_connection():
    connection = mysql.connector.connect(**MYSQL_CONFIG)
    return connection

#Random 5 alphanumeric code
def generate_random_string():
    # Define the characters to choose from
    characters = string.ascii_letters + string.digits

    # Generate a random 5-character alphanumeric string
    random_string = ''.join(random.choice(characters) for _ in range(12))

    return random_string

#Random integer
def generate_random_integer():
    return random.randint(10000, 99999)

############################################ EMAIL - EDITED TO CHECK

def send_email(sender_email, receiver_email, password, subject, msg, smtp_server, smtp_port):
    # Create a multipart message and set headers
    message = MIMEMultipart()
    message['From'] = sender_email
    message['To'] = receiver_email
    message['Subject'] = subject

    # Add body to email
    message.attach(MIMEText(msg, 'plain'))  # Changed from `message` to `body`

    # Connect to the SMTP server
    server = smtplib.SMTP(smtp_server, smtp_port)
    server.starttls()  # Upgrade the connection to TLS
    server.login(sender_email, password)

    # Send email
    server.send_message(message)

    # Close the connection
    server.quit()

def convert_to_calendar_format(date, time):
    """
    Convert date and time to the desired calendar format 'YYYYMMDDTHHMMSS'.

    Args:
        date (str): Date in the format 'YYYY-MM-DD'.
        time (str): Time in the format 'HH:MM:SS'.

    Returns:
        str: Date and time in the desired calendar format.
    """
    # Formatting the parts
    formatted_datetime = f"{date.replace('-', '')}T{time.replace(':', '')}"

    print (formatted_datetime)
    
    return formatted_datetime


def send_calendar_invitation(subject,start_time, end_time, location, description, attendees, message):
    # Email configuration
    smtp_server = 'smtp.office365.com'
    smtp_port = 587
    sender_email = 'segproject32@outlook.com'
    sender_password = 'aquastorm797'
    
    # Create message container
    msg = MIMEMultipart()
    msg['Subject'] = subject
    msg['From'] = sender_email
    msg['To'] = ', '.join(attendees)
    
    # Create calendar event content
    event_content = f'''
BEGIN:VCALENDAR
VERSION:2.0
PRODID:-//Sample Calendar//Sample Event//EN
BEGIN:VEVENT
SUMMARY:{subject}
LOCATION:{location}
DESCRIPTION:{description}
DTSTART:{start_time}
DTEND:{end_time}
END:VEVENT
END:VCALENDAR
'''
    # Attach calendar event content using MIMEText
    part = MIMEText(event_content, 'calendar')
    part.add_header('Content-Disposition', 'attachment; filename="Room Booking.ics"')
    msg.attach(part)

    # Add the message to the email body
    msg.attach(MIMEText(message, 'plain'))

    # Connect to SMTP server
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, attendees, msg.as_string())



def send_calendar_cancellation(subject, start_time, end_time, attendees, message):
    # Email configuration
    smtp_server = 'smtp.office365.com'
    smtp_port = 587
    sender_email = 'segproject32@outlook.com'
    sender_password = 'aquastorm797'
    
    # Create message container
    msg = MIMEMultipart()
    msg['Subject'] = f'{subject}'  # Update the subject to indicate cancellation
    msg['From'] = sender_email
    msg['To'] = ', '.join(attendees)
    
    # Add the message to the email body
    msg.attach(MIMEText(message, 'plain'))
    
    # Create cancellation event content
    event_content = f'''
BEGIN:VCALENDAR
VERSION:2.0
PRODID:-//Sample Calendar//Sample Event//EN
BEGIN:VEVENT
SUMMARY:CANCELED: {subject}
DTSTART:{start_time}
DTEND:{end_time}
STATUS:CANCELLED
END:VEVENT
END:VCALENDAR
'''
    # Attach cancellation event content using MIMEText
    part = MIMEText(event_content, 'calendar')
    part.add_header('Content-Disposition', 'attachment; filename="event_cancelled.ics"')
    msg.attach(part)

    # Connect to SMTP server
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, attendees, msg.as_string())

        
########################################### USER LOGIN - WORKING

# Login


class Login(BaseModel):
    user_id: str
    password: str

def user_login_check(db_connection, cursor, userID=None, password=None):
    query = "SELECT COUNT(*) FROM `user login` WHERE `User ID` = %s AND `Password` = %s"
    password_bytes = password.encode()
    password = encrypt_with_aes_ecb(password_bytes)
    cursor.execute(query, (userID, password))
    if cursor.fetchone()[0] > 0:
        send_otp(db_connection, cursor, userID)
        return True
    else:
        return False

@app.post("/login/")
def user_login(login: Login, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    if user_login_check(db_connection, cursor, userID=login.user_id, password=login.password):
        return {"message": "Login successful"}
    else:
        raise HTTPException(status_code=401, detail="Invalid credentials")


class Otp(BaseModel):
    user_id: str

def send_otp(db_connection, cursor, userID):
    # Update OTP key if login is successful
    new_otp_key = generate_random_integer()
    query = "UPDATE `user login` SET `OTPKey` = %s WHERE `User ID` = %s"
    cursor.execute(query, (new_otp_key, userID))
    otpMessage = "You're OTP for login is as follows: " + str(new_otp_key)
    send_email("segproject32@outlook.com", userID, "aquastorm797", 'Room Booking System OTP', otpMessage, 'smtp-mail.outlook.com', 587)
    db_connection.commit()  # Commit the transaction
    return True

@app.post("/SendOTP/")
def otp_control(otp: Otp, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    if send_otp(db_connection, cursor, userID=otp.user_id):
        return {"message": "OTP sent"}
    else:
        raise HTTPException(status_code=401, detail="Invalid credentials")


########################################### AUTHENTICATION - WORKING

class Auth(BaseModel):
    user_id: str
    key: int

def authentication(db_connection, cursor, userID, key):
    query = "SELECT COUNT(*) FROM `User Login` WHERE `User ID` = %s AND `OTPKey` = %s"
    cursor.execute(query, (userID, key))
    if cursor.fetchone()[0] > 0:
        # Update OTP key if login is successful
        query = "UPDATE `user login` SET `LoggedIn` = %s WHERE `User ID` = %s"
        cursor.execute(query, (1, userID))
        db_connection.commit()  # Commit the transaction
        return True
    else:
        return False
    
@app.post("/authentication/")
def user_authentication(auth: Auth, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    if authentication(db_connection, cursor, userID=auth.user_id, key=auth.key):
        return {"message": "Authentication successful"}
    else:
        raise HTTPException(status_code=401, detail="Authentication failed")

########################################### LOGOUT - WORKING

class Logout(BaseModel):
    user_id: str

def logging_out(db_connection, cursor, userID):
    try:
        # Update LoggedIn status to 0 indicating user is logged out
        query = "UPDATE `user login` SET `LoggedIn` = %s WHERE `User ID` = %s"
        cursor.execute(query, (0, userID))
        db_connection.commit()  # Commit the transaction
        return True
    except Exception as e:
        # Handle any exceptions that might occur during the process
        print(f"An error occurred: {e}")
        db_connection.rollback()  # Rollback the transaction in case of error
        return False
    
@app.post("/logout/")
def user_logout(logout: Logout, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    if logging_out(db_connection, cursor, userID=logout.user_id):
        return {"message": "User logged out successfully"}
    else:
        raise HTTPException(status_code=500, detail="Internal Server Error")
    
########################################### EDIT PASSWORD - WORKING


class EditPassword(BaseModel):
    user_id: str
    new_password: str

#error
def update_password(db_connection, cursor, userID, new_password):
    password = new_password.encode()
    new_password = encrypt_with_aes_ecb(password)
    query = "UPDATE `User Login` SET `Password` = %s WHERE `User ID` = %s"
    cursor.execute(query, (new_password, userID))
    db_connection.commit()

@app.post("/edit_password/")
def edit_user_password(edit_password: EditPassword, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    update_password(db_connection, cursor, userID=edit_password.user_id, new_password=edit_password.new_password)
    return {"message": "Password updated successfully"}


########################################### READY

class RoomAvailabilityRequest(BaseModel):
    userID: str
    capacity: int
    sec: str
    date: date
    start_time: time
    end_time: time

# Function to check room availability
def check_room_availability(cursor, userID, capacity, sec, date, start_time, end_time):
    #Get all the rooms from the sectio
    query = "SELECT `Room ID` FROM `Room` WHERE `Section` = %s;"
    cursor.execute(query, (sec,))

    rows = cursor.fetchall()

    #Get the availabilities of the different rooms (true = avaialable, false = not available)
    query1 = "SELECT COUNT(*) FROM `Booking ID description` AS BID " \
             "INNER JOIN `Booking List` AS BL ON BID.`Booking ID` = BL.`Booking ID` " \
             "WHERE BL.`Room ID` = %s AND BID.`Date` = %s " \
             "AND ((BID.`Start Time` BETWEEN %s AND %s) OR (BID.`End Time` BETWEEN %s AND %s))"

    availability_list = []      #Data stored in availability_list
    for row in rows:    
        room_id = row[0]
        cursor.execute(query1, (room_id, date, start_time, end_time, start_time, end_time))
        availability = cursor.fetchone()[0] == 0
        availability_list.append((room_id, availability))

    
    #Base color coding rooms based on available or not
    colored_availability_list = [(room_id, "green" if availability else "red") for room_id, availability in availability_list]

    #Get a list of all rooms where users can book
    rolecheck = "SELECT `role ID` from `users` where `User ID` = %s;"
    cursor.execute(rolecheck, (userID,))
    userRole = cursor.fetchall()
    role_id = userRole[0][0]    # Extracted  role

    #Only execute greyed out rooms if role is a student
    if role_id == 1:        #If student then only student access rooms will show
        query = "SELECT `Room ID` FROM `Room` WHERE `Section` = %s AND `StudentAccess` = %s ; "
        cursor.execute(query, (sec, 1))

        studentRooms = cursor.fetchall()

        student_room_ids = {room[0] for room in studentRooms}
        updated_availability_ids = {room[0] for room, _ in colored_availability_list}

        # Get the set of room IDs that are not in both lists
        ununioned_room_ids = updated_availability_ids.symmetric_difference(student_room_ids)

        # Create the final availability list using list comprehension
        final_availability_list = [
            (room_id, "grey" if room_id not in ununioned_room_ids else color)
            for room_id, color in colored_availability_list]
    else:
        final_availability_list = colored_availability_list
        

    #Change colors if the room is available but not suitable
    capacityCheck = "SELECT `Room ID` FROM `room` WHERE `Capacity` IS NOT NULL AND `Capacity` < %s;"
    cursor.execute(capacityCheck, (capacity,))
    noCapacityRooms = cursor.fetchall()
    no_capacity_room_ids = {room[0] for room in noCapacityRooms}

    final_availability_list1 = [
    (room_id, "yellow" if color == "green" and room_id in no_capacity_room_ids else color)
    for room_id, color in final_availability_list]
    
    #final print
    return final_availability_list1
    

# API endpoint to check room availability
@app.post("/check_room_availability/")
def check_room_availability_endpoint(request: RoomAvailabilityRequest, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    availability_list = check_room_availability(
        cursor,
        request.userID,
        request.capacity,
        request.sec,
        request.date,
        request.start_time.isoformat().split('.')[0],
        request.end_time.isoformat().split('.')[0]
    )
    return {"available": availability_list}




########################################### READY

class BookingRequest(BaseModel):
    user_id: str
    room_id: str
    capacity: int
    description: str
    date: date
    start_time: time
    end_time: time

def create_booking_request(db_connection, cursor, userID, room_id, capacity, description, date, start_time, end_time):

    #Has to create a new bookingRequestID which is unique.
    while True:
        # Generate a new bookingRequestID
        bookingRequestID = generate_random_string()

        # Checking if the ID is pre-existing
        bookingRequestIDCheck = "SELECT `Request ID` FROM `booking request`"
        cursor.execute(bookingRequestIDCheck)
        currentRoomIDS = cursor.fetchall()

        # Check if the bookingRequestID is in the currentRoomIDS
        if any(bookingRequestID in row for row in currentRoomIDS):
            print("Booking Request ID is in the current room IDs. Generating a new ID.")
        else:
            print("Booking Request ID is not in the current room IDs.")
            break

    #Inserting into Booking Request table first as parent table
    insert_request_query = "INSERT INTO `Booking Request` (`Request ID`, `User ID`, `Room ID`) VALUES (%s, %s, %s)"
    cursor.execute(insert_request_query, (bookingRequestID, userID, room_id))
    db_connection.commit()

    insert_description_query = "INSERT INTO `Booking Request Description` (`Request ID`, `Description`, `Date`, `Start Time`, `End Time`, `capacity`) VALUES (%s, %s, %s, %s, %s, %s)"
    cursor.execute(insert_description_query, (bookingRequestID, description, date, start_time, end_time, capacity))
    db_connection.commit()

    #Sending the confirmation email
    confirmMessage = (
        "You have sent in a booking request of ID: " +
        str(bookingRequestID) + "\n" +
        "Room ID:   " +
        str(room_id) + "\n" +
        "Booking Reason:   " +
        str(description) + "\n" +
        "Booking capacity:   " +
        str(capacity) + "\n" +
        "Date:   " +
        str(date) + "\n" +
        "Timing:   " +
        str(start_time) + "  :  " + str(end_time) + ".\nIf you did not make this booking, please check your account immediately.")

    send_email ("segproject32@outlook.com", userID, "aquastorm797",'New Booking Request', confirmMessage ,'smtp-mail.outlook.com',587 )

@app.post("/booking_request/")
def create_booking(booking_request: BookingRequest, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    create_booking_request(
        db_connection,
        cursor,
        userID=booking_request.user_id,
        room_id=booking_request.room_id,
        capacity = booking_request.capacity,
        description=booking_request.description,
        date=booking_request.date,
        start_time=booking_request.start_time.isoformat().split('.')[0],  # Convert time objects to ISO format strings
        end_time=booking_request.end_time.isoformat().split('.')[0]     # Convert time objects to ISO format strings
    )
    return {"booking successful"}




########################################### READY

class userClass(BaseModel):
    UserID: str

def get_profile_details(db_connection, cursor, userID):
    get_details = """
        SELECT `users`.`User ID`, `users`.`profile_picture`, `user roles`.`Name` 
        FROM `users` 
        JOIN `user roles` ON `users`.`Role ID` = `user roles`.`Role ID`
        WHERE `users`.`User ID` = %s
    """ 
    cursor.execute(get_details, (userID,))
    role = cursor.fetchone()

    profile = (role[0], role[1],role[2])

    return profile

@app.post("/get_profile_details/")
def get_profile(profile: userClass, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    prof =  get_profile_details(
            db_connection,
            cursor,
            userID=profile.UserID
            )
    return {"username": prof[0], "role": prof[2],"profile picture":prof[1]}



###########################################READY

#Uses userClass

def get_approving_role(db_connection, cursor, userID):
    get_details = """
        SELECT `user roles`.`Name` 
        FROM `users` 
        JOIN `user roles` ON `users`.`Role ID` = `user roles`.`Role ID`
        WHERE `users`.`User ID` = %s
    """ 
    cursor.execute(get_details, (userID,))
    
    role = cursor.fetchone()
    if role[0] in ["SAS Staff Member", "Property Manager"]:
        return True
    else:
        return False

@app.post("/check_approval_role/")
def check_approval_role(username: userClass, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    check =  get_approving_role(
            db_connection,
            cursor,
            userID=username.UserID
            )
    return (check)



########################################### READY

#Uses userClass

def get_booking_r(db_connection, cursor, userID):
    # Assuming get_profile_details is imported or defined in the same module
    profile = get_profile_details(db_connection, cursor, userID)
    role = profile[1]

    if role == "Property Manager":
        sql_query = """
            SELECT `booking request`.`Request ID`, `booking request`.`Room ID`,
                   `booking request description`.`Date`, 
                   `booking request description`.`Start Time`, 
                   `booking request description`.`End Time`
            FROM `booking request`
            JOIN `booking request description` 
                ON `booking request`.`Request ID` = `booking request description`.`Request ID`
            WHERE `booking request description`.`End Time` > '17:00:00'
        """
    else:
        sql_query = """
            SELECT `booking request`.`Request ID`, `booking request`.`Room ID`,
                   `booking request description`.`Date`, 
                   `booking request description`.`Start Time`, 
                   `booking request description`.`End Time`
            FROM `booking request`
            JOIN `booking request description` 
                ON `booking request`.`Request ID` = `booking request description`.`Request ID`
            WHERE `booking request description`.`End Time` < '17:00:00'
        """
    
    cursor.execute(sql_query)

    # Fetch all rows of the result
    result = cursor.fetchall()
     # Convert start and end times to formatted time strings
    formatted_result = []
    for row in result:
        formatted_row = list(row)
        formatted_row[3] = str(row[3])
        if formatted_row[3] == "9:00:00":
            formatted_row[3] = "09:00:00"
        formatted_row[4] = str(row[4])
        formatted_result.append(formatted_row)

    return formatted_result



@app.post("/get_booking_requests_accepter/")
def get_booking_requests(username: userClass, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    check =  get_booking_r(
            db_connection,
            cursor,
            userID=username.UserID
            )
    return check


########################################### READY


#4. ⁠get_request_details, which I will give u bookingId and need {requester: “email”, roomSpecific: {capacity: 5, purpose: [“Meeting”, Presentation”]}}
#Function name: get_request_details(BookingID)
#return: [user ID (email), user Role (student), capacity, purpose]
#syntax of {user_id: “actual email”, user_role: “student”…


class RequestDetails(BaseModel):
    bookingID: str

def get_request_details(db_connection, cursor, bookingID):
    query = """
    SELECT 
        `users`.`user ID` AS user_id,
        `user roles`.`name` AS user_role,
        `booking request description`.`capacity` AS request_capacity,
        `room`.`capacity` AS room_capacity,
        `booking request description`.`Description` AS request_description
    FROM 
        `booking request`
    JOIN 
        `users` ON `booking request`.`User ID` = `users`.`User ID`
    JOIN 
        `user roles` ON `users`.`Role ID` = `user roles`.`Role ID`
    JOIN 
        `booking request description` ON `booking request`.`Request ID` = `booking request description`.`Request ID`
    JOIN 
        `room` ON `booking request`.`Room ID` = `room`.`Room ID`
    WHERE 
        `booking request`.`Request ID` = %s

    """
    cursor.execute(query, (bookingID,))
    role = cursor.fetchone()

    return role


@app.post("/get_request_details_accepter/")
def get_request_dets(request: RequestDetails, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    details = get_request_details(db_connection, cursor, bookingID = request.bookingID)

    return {"user_id": details[0], "user_role": details[1], "request_capacity": details[2], "room_capacity": details[3], "description": details[4], }


########################################### READY


class HandleBooking(BaseModel): 
    action: str         #action is either "accept" or "decline"
    bookingID: str
    comment: str
    handler: str

def check_collision(db_connection, cursor, requestID):
    query = """
        SELECT br.`Room ID`, brd.`Date`, brd.`Start Time`, brd.`End Time`
        FROM `booking request` AS br
        JOIN `booking request description` AS brd ON br.`Request ID` = brd.`Request ID`
        WHERE br.`Request ID` = %s
    """
    cursor.execute (query, (requestID,))
    info = cursor.fetchone()

    room_id = info[0]
    date = info[1]
    start_time = info[2]
    end_time = info[3]

    #Get the availabilities of the different rooms (true = avaialable, false = not available)
    query1 = "SELECT COUNT(*) FROM `Booking ID description` AS BID " \
             "INNER JOIN `Booking List` AS BL ON BID.`Booking ID` = BL.`Booking ID` " \
             "WHERE BL.`Room ID` = %s AND BID.`Date` = %s " \
             "AND ((BID.`Start Time` BETWEEN %s AND %s) OR (BID.`End Time` BETWEEN %s AND %s))"

    cursor.execute(query1, (room_id, date, start_time, end_time, start_time, end_time))
    collision_count = cursor.fetchone()[0]
    if collision_count > 0:
                return False

    # If no collision found, return True
    return True


def accept_booking(db_connection, cursor, requestID, comment, handler):

    if check_collision(db_connection, cursor, requestID) == False:
        return False

    #Has to create a new bookingRequestID which is unique.
    while True:
        # Generate a new bookingRequestID
        NewbookingID = generate_random_string()

        # Checking if the ID is pre-existing
        bookingIDCheck = "SELECT `Booking ID` FROM `booking list`"
        cursor.execute(bookingIDCheck)
        currentBookingIDS = cursor.fetchall()

        # Check if the bookingRequestID is in the currentRoomIDS
        if any(NewbookingID in row for row in bookingIDCheck):
            print("Booking ID is in the current IDs. Generating a new ID.")     #Test print
        else:
            print("Booking ID is not in the current IDs.")                      #Test print
            break

    #New unique Booking ID named NewbookingID
    print (NewbookingID)        #Test print

    #Getting the data from the main tables for primary key purposes
    getDetails1 = ("""
        SELECT `User ID`, `Room ID`
        FROM `booking request`
        WHERE `Request ID` = %s;""")

    cursor.execute(getDetails1,(requestID,))
    details1 = cursor.fetchone()

    # Insert data into Booking List table
    insertDetails1 = ("""
        INSERT INTO `Booking List` (`Booking ID`, `User ID`, `Room ID`,`handler`)
        VALUES (%s, %s, %s, %s);""")
    cursor.execute(insertDetails1,(NewbookingID, details1[0], details1[1], handler))

    # Commit the transaction
    db_connection.commit()


    #Getting the dependant data from booking request description
    getDetails2 = ("""SELECT * FROM `booking request description` WHERE `Request ID` = %s;""")
    cursor.execute(getDetails2,(requestID,))
    details2 = cursor.fetchone()

    # Insert the data into booking id description
    insertDetails2 = ("""
        INSERT INTO `booking id description` (`Booking ID`, `Description`, `Date`, `Start Time`, `End Time` , `Capacity`, `Comment`)
        VALUES (%s, %s, %s, %s, %s, %s, %s);""")

    cursor.execute(insertDetails2,(NewbookingID, details2[1], details2[2], details2[3], details2[4], details2[5], comment))

    # Commit the transaction
    db_connection.commit()

    #Remove the data from Booking Requests as already handled

    delete1 = """
        DELETE FROM `booking request description`
        WHERE `Request ID` = %s
    """

    delete2 = """
        DELETE FROM `booking request`
        WHERE `Request ID` = %s
    """

    cursor.execute(delete1,(requestID,))
    cursor.execute(delete2,(requestID,))

    db_connection.commit()

    subject = "Booking " + str(NewbookingID)
    date = str(details2[2])
    start_time = str(details2[3])
    if start_time == "9:00:00":
        start_time = "09:00:00"
        
    end_time = str(details2[4])
    if end_time == "9:59:00":
        end_time = "09:59:00"
        
    start_time1 = convert_to_calendar_format(date, start_time)
    end_time1 = convert_to_calendar_format(date, end_time)
    
    location = details1[1]
    description = details2[1]
    attendees = [details1[0]]
    msg = "The booking request " + str(requestID) + " has been approved.\n" \
          "The ID of your booking has been updated to: " + str(NewbookingID) + "\n" \
          "Kindly add the event to your calender for reference. \n" \
        "Please view the Room Booking System Website for further details"
    
    send_calendar_invitation(subject,start_time1, end_time1, location, description, attendees, msg)
    
    return True


def decline_booking(db_connection, cursor, requestID, comment, handler):

    #Has to create a new Reject ID which is unique.
    while True:
        # Generate a new bookingRequestID
        NewbookingID = generate_random_string()

        # Checking if the ID is pre-existing
        RejectIDCheck = "SELECT `Reject ID` FROM `booking rejects`"
        cursor.execute(RejectIDCheck)
        currentRejectIDS = cursor.fetchall()

        # Check if the bookingRequestID is in the currentRoomIDS
        if any(NewbookingID in row for row in currentRejectIDS):
            print("Booking ID is in the current IDs. Generating a new ID.")     #Test print
        else:
            print("Booking ID is not in the current IDs.")                      #Test print
            break

    #New unique Booking ID named NewbookingID
    print (NewbookingID)        #Test print

    #Getting the data from the main tables for primary key purposes
    getDetails1 = ("""
        SELECT `User ID`, `Room ID`
        FROM `booking request`
        WHERE `Request ID` = %s;""")

    cursor.execute(getDetails1,(requestID,))
    details1 = cursor.fetchone()

    # Insert data into Booking List table
    insertDetails1 = ("""
        INSERT INTO `booking rejects` (`Reject ID`, `User ID`, `Room ID`,`handler`)
        VALUES (%s, %s, %s, %s);""")
    cursor.execute(insertDetails1,(NewbookingID, details1[0], details1[1], handler))

    # Commit the transaction
    db_connection.commit()


    #Getting the dependant data from booking request description
    getDetails2 = ("""SELECT * FROM `booking request description` WHERE `Request ID` = %s;""")
    cursor.execute(getDetails2,(requestID,))
    details2 = cursor.fetchone()

    # Insert the data into booking id description
    insertDetails2 = ("""
        INSERT INTO `booking rejects description` (`Reject ID`, `Description`, `Date`, `Start Time`, `End Time` , `Capacity`, `Comment`)
        VALUES (%s, %s, %s, %s, %s, %s, %s);""")

    cursor.execute(insertDetails2,(NewbookingID, details2[1], details2[2], details2[3], details2[4], details2[5], comment))

    #Remove the data from Booking Requests as already handled
    delete1 = """
        DELETE FROM `booking request description`
        WHERE `Request ID` = %s
    """

    delete2 = """
        DELETE FROM `booking request`
        WHERE `Request ID` = %s
    """

    cursor.execute(delete1,(requestID,))
    cursor.execute(delete2,(requestID,))

    db_connection.commit()
    
    msg = "The booking request " + str(requestID) + " has been declined.\n" \
          "The booking is now rejected, with ID: " + str(NewbookingID) + ". \n" \
          "The reason for the booking declined is: \n" + comment + "\n" \
          "Please view the Room Booking System Website for further details"
    send_email ("segproject32@outlook.com", details1[0], "aquastorm797",'Room Booking Request Declined', msg ,'smtp-mail.outlook.com',587 )



@app.post("/handle_booking/")
def handle_booking(handling: HandleBooking, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    
    action = handling.action

    if action == "approve":
        if accept_booking(db_connection, cursor, handling.bookingID, handling.comment, handling.handler) == True:
            return {"Approval Complete"}
        else:
            return {"Collision"}
    else:
        decline_booking(db_connection, cursor, handling.bookingID, handling.comment, handling.handler)
        return {"Decline complete"}





#################################### READY
    
class HandleRequests(BaseModel):
    UserID: str
    checkType: str      #action is either "current" or "past"


#Supposed to take userID and return the current or past requests based on todays date. Current is > todays date, past is < todays date
def get_user_requests(db_connection, cursor, UserID, checkType):
    import datetime
    #Get todays date for checking
    today = datetime.date.today()
    formatted_date = today.strftime("%Y-%m-%d")

    print (formatted_date)

    if checkType == "current":
        sql_query1 = """
            SELECT `booking request`.`Request ID`, `booking request`.`Room ID`,
                   `booking request description`.`Date`,
                   `booking request description`.`Start Time`,
                   `booking request description`.`End Time`
            FROM `booking request`
            JOIN `booking request description`
            ON `booking request`.`Request ID` = `booking request description`.`Request ID`
            WHERE `booking request description`.`Date` >= %s
            AND `booking request`.`User ID` = %s
        """

        sql_query2 = """
            SELECT `booking list`.`Booking ID`, `booking list`.`Room ID`,
                   `booking id description`.`Date`,
                   `booking id description`.`Start Time`,
                   `booking id description`.`End Time`
            FROM `booking list`
            JOIN `booking id description`
            ON `booking list`.`Booking ID` = `booking id description`.`Booking ID`
            WHERE `booking id description`.`Date` >= %s
            AND `booking list`.`User ID` = %s
        """

        sql_query3 = """
            SELECT `booking rejects`.`Reject ID`, `booking rejects`.`Room ID`,
                   `booking rejects description`.`Date`,
                   `booking rejects description`.`Start Time`,
                   `booking rejects description`.`End Time`
            FROM `booking rejects`
            JOIN `booking rejects description`
            ON `booking rejects`.`Reject ID` = `booking rejects description`.`Reject ID`
            WHERE `booking rejects description`.`Date` >= %s
            AND `booking rejects`.`User ID` = %s
        """

        cursor.execute(sql_query1, (formatted_date,UserID))
        results1 = cursor.fetchall()
        results1_labeled = [row + ("Pending",) for row in results1]

        cursor.execute(sql_query2, (formatted_date,UserID))
        results2 = cursor.fetchall()
        results2_labeled = [row + ("Accepted",)  for row in results2]

        cursor.execute(sql_query3, (formatted_date,UserID))
        results3 = cursor.fetchall()
        results3_labeled = [row + ("Rejected",)  for row in results3]    

    else:
        sql_query1 = """
            SELECT `booking request`.`Request ID`, `booking request`.`Room ID`,
                   `booking request description`.`Date`,
                   `booking request description`.`Start Time`,
                   `booking request description`.`End Time`
            FROM `booking request`
            JOIN `booking request description`
            ON `booking request`.`Request ID` = `booking request description`.`Request ID`
            WHERE `booking request description`.`Date` < %s
            AND `booking request`.`User ID` = %s
        """

        sql_query2 = """
            SELECT `booking list`.`Booking ID`, `booking list`.`Room ID`,
                   `booking id description`.`Date`,
                   `booking id description`.`Start Time`,
                   `booking id description`.`End Time`
            FROM `booking list`
            JOIN `booking id description`
            ON `booking list`.`Booking ID` = `booking id description`.`Booking ID`
            WHERE `booking id description`.`Date` < %s
            AND `booking list`.`User ID` = %s
        """

        sql_query3 = """
            SELECT `booking rejects`.`Reject ID`, `booking rejects`.`Room ID`,
                   `booking rejects description`.`Date`,
                   `booking rejects description`.`Start Time`,
                   `booking rejects description`.`End Time`
            FROM `booking rejects`
            JOIN `booking rejects description`
            ON `booking rejects`.`Reject ID` = `booking rejects description`.`Reject ID`
            WHERE `booking rejects description`.`Date` < %s
            AND `booking rejects`.`User ID` = %s
        """        

        cursor.execute(sql_query1, (formatted_date,UserID))
        results1 = cursor.fetchall()
        results1_labeled = [row + ("Pending",) for row in results1]

        cursor.execute(sql_query2, (formatted_date,UserID))
        results2 = cursor.fetchall()
        results2_labeled = [row + ("Completed",)  for row in results2]

        cursor.execute(sql_query3, (formatted_date,UserID))
        results3 = cursor.fetchall()
        results3_labeled = [row + ("Rejected",)  for row in results3]    


    results = results1_labeled + results2_labeled + results3_labeled

    formatted_result = []

    for row in results:
        formatted_row = list(row)
        formatted_row[3] = str(row[3])
        if formatted_row[3] == "9:00:00":
            formatted_row[3] = "09:00:00"
        formatted_row[4] = str(row[4])
        formatted_result.append(formatted_row)

    return formatted_result

@app.post("/get_booking_requests_users/")
def get_booking_requests_users(handling: HandleRequests, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    check = get_user_requests(
            db_connection,
            cursor,
            UserID=handling.UserID,
            checkType=handling.checkType
            )
    return check




####################################READY


class BookingDetailsUsers(BaseModel):
    ID: str
    checkType: str      #action is either "Approved" OR "Completed", OR "Rejected" OR "Pending"


def get_user_booking_details(db_connection, cursor, ID, checkType):

    if checkType == "Pending":  #Pending
        
        query = """
            SELECT 
                `booking request description`.`capacity` AS request_capacity,
                `room`.`capacity` AS room_capacity,
                `booking request description`.`Description` AS request_description
            FROM 
                `booking request`
            JOIN 
                `booking request description` ON `booking request`.`Request ID` = `booking request description`.`Request ID`
            JOIN 
                `room` ON `booking request`.`Room ID` = `room`.`Room ID`
            WHERE 
                `booking request`.`Request ID` = %s
            """

    elif checkType == "Approved" or checkType == "Completed":   #Approved or Completed
        query = """
            SELECT 
                `booking id description`.`capacity` AS request_capacity,
                `room`.`capacity` AS room_capacity,
                `booking id description`.`Description` AS request_description,
                `booking id description`.`Comment` AS request_comment
            FROM 
                `booking list`
            JOIN 
                `booking id description` ON `booking list`.`Booking ID` = `booking id description`.`Booking ID`
            JOIN 
                `room` ON `booking list`.`Room ID` = `room`.`Room ID`
            WHERE 
                `booking list`.`Booking ID` = %s
        """

        checkComment = "SELECT `Title`, `Text` FROM `feedback` WHERE `Booking ID` = %s"
        cursor.execute(checkComment,(ID,))
        validID = cursor.fetchone()

    else:   #Rejected
        
        query = """
            SELECT 
                `booking rejects description`.`capacity` AS request_capacity,
                `room`.`capacity` AS room_capacity,
                `booking rejects description`.`Description` AS request_description,
                `booking rejects description`.`Comment` AS request_comment
            FROM 
                `booking rejects`
            JOIN 
                `booking rejects description` ON `booking rejects`.`Reject ID` = `booking rejects description`.`Reject ID`
            JOIN 
                `room` ON `booking rejects`.`Room ID` = `room`.`Room ID`
            WHERE 
                `booking rejects`.`Reject ID` = %s
            """

        
    
    cursor.execute(query, (ID,))
    results = cursor.fetchone()

    if  checkType == "Approved" or checkType == "Completed":
        if validID:
            response = {
                    "request_capacity": results[0] if results[0] is not None else "",
                    "room_capacity": results[1] if results[1] is not None else "",
                    "description": results[2] if results[2] is not None else "",
                    "comment": results[3] if results[3] is not None else "",
                    "feedback_title": validID[0] if validID[0] is not None else "",
                    "feedback_text": validID[1] if validID[1] is not None else "",
                }

        else:
            
            response = {
                    "request_capacity": results[0] if results[0] is not None else "",
                    "room_capacity": results[1] if results[1] is not None else "",
                    "description": results[2] if results[2] is not None else "",
                    "comment": results[3] if results[3] is not None else "",
                }


    elif checkType == "Rejected":
        response = {
                "request_capacity": results[0] if results[0] is not None else "",
                "room_capacity": results[1] if results[1] is not None else "",
                "description": results[2] if results[2] is not None else "",
                "comment": results[3] if results[3] is not None else "",
            }
        
    else:   
        response = {
            "request_capacity": results[0] if results[0] is not None else "",
            "room_capacity": results[1] if results[1] is not None else "",
            "description": results[2] if results[2] is not None else "",
        }


    return response
    
#need room capacity, booking capacity, purpose, comment
#        `booking id description`.`


@app.post("/get_booking_details_users/")
def get_booking_details_users(details: BookingDetailsUsers, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    check = get_user_booking_details(
            db_connection,
            cursor,
            ID=details.ID,
            checkType=details.checkType
            )
    return check



#################################### AUTOMATION CALL

#call everyday at 00:00:00 for automation of clear bookings past current date
def daily_bookingsR_clear(db_connection, cursor):
    # Get todays date and save locally for the clear
    today = datetime.now().date()

    # Calculate one day prior
    one_day_prior = today - timedelta(days=1)
    one_day_prior_str = one_day_prior.strftime('%Y-%m-%d')
    
    query = "SELECT `Request ID` FROM `booking request description` WHERE `Date` < %s"
    cursor.execute(query, (one_day_prior_str,))
    bookings_today = cursor.fetchall()  
    print (bookings_today)
    
    if bookings_today is not None:
        #Loop through and decline it 
        for request_id in bookings_today:
            decline_booking(db_connection, cursor, request_id[0], "Booking declined automatically as it's past the scheduled date.","System")

@app.post("/daily_bookingsR_clear/")
def get_booking_requests_users(db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    daily_bookingsR_clear(db_connection,cursor)
    return {"Complete clearing"}

#Alt 3 for comment, Alt 4 for uncomment


#################################### READY 

class bookingType(BaseModel):
    typeCheck : str


def getAllBookings (db_connection, cursor,typeCheck):
    import datetime
    #Get todays date for checking
    today = datetime.date.today()
    formatted_date = today.strftime("%Y-%m-%d")

    print (formatted_date)

    if typeCheck == "current":
        sql_query1 = """
            SELECT `booking request`.`Request ID`, `booking request`.`Room ID`,
                   `booking request description`.`Date`,
                   `booking request description`.`Start Time`,
                   `booking request description`.`End Time`
            FROM `booking request`
            JOIN `booking request description`
            ON `booking request`.`Request ID` = `booking request description`.`Request ID`
            WHERE `booking request description`.`Date` >= %s
        """

        sql_query2 = """
            SELECT `booking list`.`Booking ID`, `booking list`.`Room ID`,
                   `booking id description`.`Date`,
                   `booking id description`.`Start Time`,
                   `booking id description`.`End Time`
            FROM `booking list`
            JOIN `booking id description`
            ON `booking list`.`Booking ID` = `booking id description`.`Booking ID`
            WHERE `booking id description`.`Date` >= %s
        """

        sql_query3 = """
            SELECT `booking rejects`.`Reject ID`, `booking rejects`.`Room ID`,
                   `booking rejects description`.`Date`,
                   `booking rejects description`.`Start Time`,
                   `booking rejects description`.`End Time`
            FROM `booking rejects`
            JOIN `booking rejects description`
            ON `booking rejects`.`Reject ID` = `booking rejects description`.`Reject ID`
            WHERE `booking rejects description`.`Date` >= %s
        """

        cursor.execute(sql_query1, (formatted_date,))
        results1 = cursor.fetchall()
        results1_labeled = [row + ("Pending",) for row in results1]

        cursor.execute(sql_query2, (formatted_date,))
        results2 = cursor.fetchall()
        results2_labeled = [row + ("Accepted",)  for row in results2]

        cursor.execute(sql_query3, (formatted_date,))
        results3 = cursor.fetchall()
        results3_labeled = [row + ("Rejected",)  for row in results3]    

    else:
        sql_query1 = """
            SELECT `booking request`.`Request ID`, `booking request`.`Room ID`,
                   `booking request description`.`Date`,
                   `booking request description`.`Start Time`,
                   `booking request description`.`End Time`
            FROM `booking request`
            JOIN `booking request description`
            ON `booking request`.`Request ID` = `booking request description`.`Request ID`
            WHERE `booking request description`.`Date` < %s
        """

        sql_query2 = """
            SELECT `booking list`.`Booking ID`, `booking list`.`Room ID`,
                   `booking id description`.`Date`,
                   `booking id description`.`Start Time`,
                   `booking id description`.`End Time`
            FROM `booking list`
            JOIN `booking id description`
            ON `booking list`.`Booking ID` = `booking id description`.`Booking ID`
            WHERE `booking id description`.`Date` < %s
        """

        sql_query3 = """
            SELECT `booking rejects`.`Reject ID`, `booking rejects`.`Room ID`,
                   `booking rejects description`.`Date`,
                   `booking rejects description`.`Start Time`,
                   `booking rejects description`.`End Time`
            FROM `booking rejects`
            JOIN `booking rejects description`
            ON `booking rejects`.`Reject ID` = `booking rejects description`.`Reject ID`
            WHERE `booking rejects description`.`Date` < %s
        """        

        cursor.execute(sql_query1, (formatted_date,))
        results1 = cursor.fetchall()
        results1_labeled = [row + ("Pending",) for row in results1]

        cursor.execute(sql_query2, (formatted_date,))
        results2 = cursor.fetchall()
        results2_labeled = [row + ("Completed",)  for row in results2]

        cursor.execute(sql_query3, (formatted_date,))
        results3 = cursor.fetchall()
        results3_labeled = [row + ("Rejected",)  for row in results3]    


    results = results1_labeled + results2_labeled + results3_labeled

    formatted_result = []

    for row in results:
        formatted_row = list(row)
        formatted_row[3] = str(row[3])
        if formatted_row[3] == "9:00:00":
            formatted_row[3] = "09:00:00"
        formatted_row[4] = str(row[4])
        formatted_result.append(formatted_row)

    return formatted_result


@app.post("/get_All_Bookings_Admin/")
def get_booking_requests_admin(bookingtype : bookingType, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    dataset = getAllBookings(db_connection,cursor, typeCheck = bookingtype.typeCheck)
    return dataset

#################################### READY 


class booking(BaseModel):
    bookingID : str

def get_admin_details(db_connection,cursor, ID):
    query = """
        SELECT
            `booking list`.`User ID`,
            `user roles`.`Name`,
            `booking list`.`Handler`,
            `booking id description`.`capacity` AS request_capacity,
            `room`.`capacity` AS room_capacity,
            `booking id description`.`Description` AS request_description,
            `booking id description`.`Comment` AS request_comment
        FROM 
            `booking list`
        JOIN 
            `users` ON `booking list`.`User ID` = `users`.`User ID`
        JOIN 
            `user roles` ON `users`.`Role ID` = `user roles`.`Role ID`
        JOIN 
            `booking id description` ON `booking list`.`Booking ID` = `booking id description`.`Booking ID`
        JOIN 
            `room` ON `booking list`.`Room ID` = `room`.`Room ID`
            
        WHERE 
            `booking list`.`Booking ID` = %s
         """
    query1 = """
        SELECT
            `booking rejects`.`User ID`,
            `user roles`.`Name`,
            `booking rejects`.`Handler`,
            `booking rejects description`.`capacity` AS request_capacity,
            `room`.`capacity` AS room_capacity,
            `booking rejects description`.`Description` AS request_description,
            `booking rejects description`.`Comment` AS request_comment
        FROM 
            `booking rejects`
        JOIN 
            `users` ON `booking rejects`.`User ID` = `users`.`User ID`
        JOIN 
            `user roles` ON `users`.`Role ID` = `user roles`.`Role ID`
        JOIN 
            `booking rejects description` ON `booking rejects`.`Reject ID` = `booking rejects description`.`Reject ID`
        JOIN 
            `room` ON `booking rejects`.`Room ID` = `room`.`Room ID`
            
        WHERE 
            `booking rejects`.`Reject ID` = %s
         """

    cursor.execute(query, (ID,))
    details = cursor.fetchone()

    if details == None:
        cursor.execute(query1, (ID,))
        details = cursor.fetchone()


    transformed_data = {
    "user_id": details[0],
    "user_role": details[1],
    "handler": details[2],
    "request_capacity": details[3],
    "room_capacity": details[4],
    "description": details[5],
    "comment": details[6]
}

    return transformed_data



@app.post("/get_Booking_Details_Admin/")
def get_Booking_Details_Admin(booking : booking, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    dataset = get_admin_details(db_connection,cursor, ID = booking.bookingID)
    return dataset

#################################### READY

class cancelations(BaseModel):
    bookingID : str
    reason : str
    handler : str

def cancelBooking (db_connection,cursor, ID, reason, handler):
    #Has to create a new Reject ID which is unique.
    while True:
        # Generate a new bookingRequestID
        NewbookingID = generate_random_string()

        # Checking if the ID is pre-existing
        RejectIDCheck = "SELECT `Reject ID` FROM `booking rejects`"
        cursor.execute(RejectIDCheck)
        currentRejectIDS = cursor.fetchall()

        # Check if the bookingRequestID is in the currentRoomIDS
        if any(NewbookingID in row for row in currentRejectIDS):
            print("Booking ID is in the current IDs. Generating a new ID.")     #Test print
        else:
            print("Booking ID is not in the current IDs.")                      #Test print
            break

    #New unique Booking ID named NewbookingID
    print (NewbookingID)        #Test print

    #Getting the data from the main tables for primary key purposes
    getDetails1 = ("""
        SELECT `User ID`, `Room ID`
        FROM `booking list`
        WHERE `Booking ID` = %s;""")

    cursor.execute(getDetails1,(ID,))
    details1 = cursor.fetchone()

    # Insert data into Booking List table
    insertDetails1 = ("""
        INSERT INTO `booking rejects` (`Reject ID`, `User ID`, `Room ID`, `handler`)
        VALUES (%s, %s, %s, %s);""")
    cursor.execute(insertDetails1,(NewbookingID, details1[0], details1[1], handler))

    # Commit the transaction
    db_connection.commit()

    #Getting the dependant data from booking request description
    getDetails2 = ("""SELECT * FROM `booking id description` WHERE `Booking ID` = %s;""")
    cursor.execute(getDetails2,(ID,))
    details2 = cursor.fetchone()

    # Insert the data into booking id description
    insertDetails2 = ("""
        INSERT INTO `booking rejects description` (`Reject ID`, `Description`, `Date`, `Start Time`, `End Time` , `Capacity`, `Comment`)
        VALUES (%s, %s, %s, %s, %s, %s, %s);""")

    cursor.execute(insertDetails2,(NewbookingID, details2[1], details2[2], details2[3], details2[4], details2[5], reason))

    #Remove the data from Booking Requests as already handled
    delete1 = """
        DELETE FROM `booking id description`
        WHERE `Booking ID` = %s
    """

    delete2 = """
        DELETE FROM `booking list`
        WHERE `Booking ID` = %s
    """

    cursor.execute(delete1,(ID,))
    cursor.execute(delete2,(ID,))

    db_connection.commit()
    
    subject = "Booking " + str(ID)
    date = str(details2[2])
    
    start_time = str(details2[3])
    if start_time == "9:00:00":
        start_time = "09:00:00"
        
    end_time = str(details2[4])
    if end_time == "9:59:00":
        end_time = "09:59:00"

    start_time1 = str(convert_to_calendar_format(date, start_time))
    end_time1 =  str(convert_to_calendar_format(date, end_time))
    
    attendees = [details1[0]]
    message = "The booking ID " + str(ID) + " has now been canceled.\n" \
          "The booking shows as rejected, with ID: " + str(NewbookingID) + ". \n" \
          "Please view the Room Booking System Website for further details"    

    send_calendar_cancellation(subject, start_time1, end_time1, attendees, message)

@app.post("/booking_Cancel/")
def cancel_booking(cancel : cancelations, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    cancelBooking(db_connection,cursor, ID = cancel.bookingID, reason = cancel.reason, handler = cancel.handler)
    return {"Complete"}



#################################### READY

class room(BaseModel):
    roomID : str

def get_room_quantity(db_connection, cursor, ID):
    query1 = "SELECT `Capacity` FROM `room` WHERE `Room ID` = %s"
    
    cursor.execute(query1,(ID,))

    details = cursor.fetchone()

    quantity = details[0]

    return quantity
    
def get_room_details(db_connection, cursor, ID):
    query1 = "SELECT `equipment`, `quantity` FROM item WHERE `Room ID` = %s"

    cursor.execute(query1,(ID,))

    details = cursor.fetchall()

    return details

@app.post("/Room_Details/")
def room_Details(room : room, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    quantity = get_room_quantity(db_connection,cursor, ID = room.roomID)
    details = get_room_details(db_connection,cursor, ID = room.roomID)
    return details

@app.post("/Room_Quantity/")
def room_Quantity(room : room, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    quantity = get_room_quantity(db_connection,cursor, ID = room.roomID)
    return quantity


#################################### READY

#################################### FEEDBACK SECTION

class Feedback(BaseModel):
    bookingID : str
    title : str
    feedback : str

def createFeedback(db_connection, cursor, bookingID, title, feedback):
    # Check if the bookingID already exists
    query_check = "SELECT COUNT(*) FROM feedback WHERE `Booking ID` = %s"
    cursor.execute(query_check, (bookingID,))
    result = cursor.fetchone()

    if result[0] > 0:  # If bookingID exists, update the existing record
        query_update = """
            UPDATE feedback 
            SET `Title` = %s, `Text` = %s, `Active` = %s 
            WHERE `Booking ID` = %s
        """
        cursor.execute(query_update, (title, feedback, True, bookingID))
    else:  # If bookingID does not exist, insert a new record
        query_insert = """
            INSERT INTO feedback (`Booking ID`, `Title`, `Text`, `Active`)
            VALUES (%s, %s, %s, %s)
        """
        cursor.execute(query_insert, (bookingID, title, feedback, True))

    db_connection.commit()

    

@app.post("/create_Feedback/")
def create_Feedback(fb : Feedback, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    createFeedback(db_connection,cursor, bookingID = fb.bookingID, title = fb.title, feedback = fb.feedback)
    return {"Complete"}


#class booking(BaseModel):
#    bookingID : str

def readFeedback(db_connection,cursor, ID):

    read_query = "SELECT `Active` FROM `feedback` WHERE `Booking ID` = %s"
    cursor.execute(read_query,(ID,))
    value = cursor.fetchone()
    value1 = value[0]

    value2 = not value1
    
    update_query = """
        UPDATE feedback
        SET `Active` = %s
        WHERE `Booking ID` = %s
    """

    cursor.execute(update_query,(value2,ID))

    # Commit the transaction
    db_connection.commit()


@app.post("/read_Feedback/")
def read_Feedback(bt : booking, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    readFeedback(db_connection,cursor, ID = bt.bookingID)
    return {"Complete"}


#class bookingType(BaseModel):
    #typeCheck : str

def get_list_feedback(db_connection,cursor, typeCheck):
    if typeCheck == "current":
        query = """
            SELECT `Booking ID`, `Title`, `Text`
            FROM `feedback`
            WHERE `Active` = 1
        """

    else:
        query = """
            SELECT `Booking ID`, `Title`, `Text`
            FROM `feedback`
            WHERE `Active` = 0
        """

    cursor.execute(query)
    feedbacklist = cursor.fetchall()

    return feedbacklist


@app.post("/get_Feedback/")             #"current" or "past"
def get_Feedback(bt : bookingType, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    feedbacklist = get_list_feedback(db_connection,cursor, typeCheck = bt.typeCheck)
    return feedbacklist


#################################### READY


def update_profile_picture_in_db(db_connection, cursor, user_id, image_data):
    # Update the profile picture in the database
    update_query = "UPDATE users SET profile_picture = %s WHERE `User ID` = %s"
    cursor.execute(update_query, (image_data, user_id))
    db_connection.commit()

@app.post("/update_profile_pic/")
def update_profile_picture(user_id: str, file: UploadFile = UploadFile(...), db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):

    if file.filename == "":
        # Handle the case when no file is uploaded
        raise HTTPException(status_code=400, detail="No file uploaded")
    else:
        cursor = db_connection.cursor()
        
        # Read the file content as bytes
        profile_picture_bytes = file.file.read()
        
        update_profile_picture_in_db(db_connection, cursor, user_id, profile_picture_bytes)
    
        return {"message": "Profile picture updated successfully"}


#################################### READY

def insert_data(db_connection, cursor, userlist):
    insert_user_query = "INSERT INTO `users` (`User ID`, `Role ID`) VALUES (%s, %s)"
    insert_user_login_query = "INSERT INTO `user login` (`User ID`, Password) VALUES (%s, %s)"

    user_check = "SELECT * FROM `users` WHERE `User ID` = %s"
    
    for user in userlist:
        cursor.execute(user_check,(user[0],))
        if cursor.fetchall():
            next
        else:  
            temp = str(user[1]).encode()
            password = encrypt_with_aes_ecb(temp)
                
            cursor.execute(insert_user_query, (user[0], user[2]))
            cursor.execute(insert_user_login_query, (user[0], password))
    # Commit the changes to the database
    db_connection.commit()

@app.post("/upload-excel/")
async def upload_excel(file: UploadFile = File(...), db_connection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    # Check if the file is an Excel file
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx) are supported.")

    # Save the uploaded Excel file to a temporary file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(await file.read())
        tmp_filename = tmp.name

    # Read the Excel file using pandas
    excel_data = pd.read_excel(tmp_filename)

    # Iterate over the rows of the Excel file
    for index, row in excel_data.iterrows():
        # Extract data from each row
        username = row['username']
        password = row['password']
        user_role = row['user role']

        # Insert the user data into the database
        insert_data(db_connection, cursor, [(username, password, user_role)])

    return {"message": "Excel upload complete"}



#################################### TO BE COMPLETED 

class graphdata(BaseModel):         #If its data for May 2024, then should get "05" "2024"
    GraphType: str  #Refer below to the name of graphs
    Month : int
    Year : int 

def format_result(result):
    labels = []
    data = []

    for row in result:
        labels.append(row[0])
        data.append(row[1])

    return {"labels": labels, "data": data}
    
#Number of rooms booked per floor
def graphFloor(cursor, db_connection, month, year):
    query = """SELECT room.Section, COUNT(*) AS occurrence_count
            FROM room
            JOIN `booking list` ON room.`Room ID` = `booking list`.`Room ID`
            JOIN `booking id description` ON `booking list`.`Booking ID` = `booking id description`.`Booking ID`
            WHERE YEAR(`booking id description`.`Date`) = %s
              AND MONTH(`booking id description`.`Date`) = %s
            GROUP BY room.Section;
                        """
        
    cursor.execute(query, (year, month,))
    formatted = format_result(cursor.fetchall())

    return formatted
    

#Number of bookings per role category
def graphRole(cursor, db_connection, month, year):

    query = """
        SELECT `user roles`.`Name`, COUNT(*) AS occurrence_count
        FROM `user roles`
        JOIN `users` ON `user roles`.`Role ID` = `users`.`Role ID`  
        JOIN `booking list` ON `users`.`User ID` = `booking list`.`User ID`       
        JOIN `booking id description` ON `booking list`.`Booking ID` = `booking id description`.`Booking ID`
        WHERE YEAR(`booking id description`.`Date`) = %s
          AND MONTH(`booking id description`.`Date`) = %s
        GROUP BY `user roles`.`Name`;
    """

    cursor.execute(query, (year, month,))
    formatted = format_result(cursor.fetchall())

    return formatted
    

    
#Capacity per bookings (group by 10s)
def graphCapacity(cursor, db_connection, month, year):

    query = """
        SELECT CONCAT(FLOOR(`capacity` / 10) * 10, ' - ', FLOOR(`capacity` / 10) * 10 + 9) AS `Capacity Group`, COUNT(*) AS group_count
        FROM `booking id description`
        WHERE YEAR(`booking id description`.`Date`) = %s
          AND MONTH(`booking id description`.`Date`) = %s
        GROUP BY FLOOR(`capacity` / 10);
    """

    cursor.execute(query, (year, month,))
    formatted = format_result(cursor.fetchall())

    return formatted
    
    
#Most booked rooms (top 10)
def graphMostBooked(cursor, db_connection, month, year):
    query = """
        SELECT `Room ID`, COUNT(*) AS group_count
        FROM `booking list`
        JOIN `booking id description` ON `booking list`.`Booking ID` = `booking id description`.`Booking ID`
        WHERE YEAR(`booking id description`.`Date`) = %s
          AND MONTH(`booking id description`.`Date`) = %s
        GROUP BY `Room ID`
        ORDER BY group_count DESC
        LIMIT 10;
    """

    cursor.execute(query, (year, month,))
    formatted = format_result(cursor.fetchall())

    return formatted
    
#Least booked rooms (top 10)
def graphLeastBooked(cursor, db_connection, month, year):
    query = """
        SELECT `Room ID`, COUNT(*) AS group_count
        FROM `booking list`
        JOIN `booking id description` ON `booking list`.`Booking ID` = `booking id description`.`Booking ID`
        WHERE YEAR(`booking id description`.`Date`) = %s
          AND MONTH(`booking id description`.`Date`) = %s
        GROUP BY `Room ID`
        ORDER BY group_count ASC
        LIMIT 10;
    """

    cursor.execute(query, (year, month,))
    formatted = format_result(cursor.fetchall())

    return formatted
   
#Which admin accepts bookings
def graphAdmins(cursor, db_connection, month, year):
    query = """
        SELECT `handler`, COUNT(*) AS group_count
        FROM `booking list`
        JOIN `booking id description` ON `booking list`.`Booking ID` = `booking id description`.`Booking ID`
        WHERE YEAR(`booking id description`.`Date`) = %s
          AND MONTH(`booking id description`.`Date`) = %s
        GROUP BY `handler`;
    """

    cursor.execute(query, (year, month,))
    formatted = format_result(cursor.fetchall())

    return formatted

#Bookings based on time
def graphTimes(cursor, db_connection, month, year):

    dataset = []
    for hour in range(9, 23):
        time_str = f"{hour:02d}:00:00"  # Format the hour as two digits
        dataset.append(time_str)

    print(dataset)
               
    query = """
        SELECT COUNT(*) as bookings
        FROM `booking id description`
        WHERE %s BETWEEN `Start Time` AND `End Time`
          AND YEAR(`booking id description`.`Date`) = %s
          AND MONTH(`booking id description`.`Date`) = %s;
    """

    labels = []
    data = []

    for time in dataset:
        cursor.execute(query, (time,year, month,))
        counter = cursor.fetchone()
        labels.append(time)
        data.append(counter[0])
        

    return {"labels": labels, "data": data}


@app.post("/get_graph/")
def upload_excel(gd : graphdata, db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    
    cursor = db_connection.cursor()
    graphtype = gd.GraphType
    month = gd.Month
    year = gd.Year

    #REFER TO THIS IF ELSE TO SEE WHAT THE GRAPHTYPES ARE
    if graphtype == 'floor':
        return graphFloor(cursor, db_connection, month, year)
            
    elif graphtype == 'role':
        return graphRole(cursor, db_connection, month, year)
        
    elif graphtype == 'capacity':
        return graphCapacity(cursor, db_connection, month, year)
        
    elif graphtype == 'most':
        return graphMostBooked(cursor, db_connection, month, year)
        
    elif graphtype == 'least':
        return graphLeastBooked(cursor, db_connection, month, year)
        
    elif graphtype == 'admins':
        return graphAdmins(cursor, db_connection, month, year)
        
    elif graphtype == 'times':
        return graphTimes(cursor, db_connection, month, year)

######################################################

# Function to get the value from the nearest non-empty cell above the specified cell in the same column
async def get_value_above(worksheet, cell):
    if not cell.value:
        current_row = cell.row
        current_column = cell.column
        while current_row > 1:
            current_row -= 1
            above_cell = worksheet.cell(row=current_row, column=current_column)
            if above_cell.value:
                return above_cell.value
    return cell.value

async def get_value_left(worksheet, cell):
    if not cell.value:
        current_row = cell.row
        current_column = cell.column
        while current_column > 1:
            current_column -= 1
            left_cell = worksheet.cell(row=current_row, column=current_column)
            if left_cell.value:
                return left_cell.value
    return cell.value

async def process_excel(file: UploadFile):
    # Load the workbook from the file contents
    file_contents = await file.read()  # Await the coroutine to get the file contents
    workbook = openpyxl.load_workbook(filename=BytesIO(file_contents))
    worksheet = workbook.active

    start_row = 5
    start_column = 2
    totaldata = []

    for column_index in range(start_column, worksheet.max_column + 1):
        for row_index in range(start_row,worksheet.max_row + 1, 3):
            classroom = await get_value_above(worksheet, worksheet.cell(row=row_index, column=1))
            day = await get_value_left(worksheet, worksheet.cell(row=3, column=column_index))
            cell_b11_value = worksheet.cell(row=row_index, column=column_index).value
            cell_b12_value = worksheet.cell(row=row_index + 1, column=column_index).value
            cell_b13_value = worksheet.cell(row=row_index + 2, column=column_index).value
            combined_values = ','.join(str(cell_value) for cell_value in [cell_b11_value, cell_b12_value, cell_b13_value] if cell_value is not None)

            if combined_values:
                dataset = f"[{classroom},{day}, {combined_values}]"
                dataset = dataset.replace(" ", "")


                if dataset.count(',') == 4:
                    parts = dataset.strip("[]").split(",")
                    time_range = parts[3]
                    start_str, end_str = time_range.split("-")
                    start_str = start_str.strip()
                    end_str = end_str.strip()
                    start_time = datetime.strptime(start_str, "%I:%M%p")
                    end_time = datetime.strptime(end_str, "%I:%M%p")
                    stime_only = start_time.time()
                    etime_only = end_time.time()
                    output_string = f"{parts[0]}, {parts[1]}, {parts[2]}, {stime_only}, {etime_only}, {parts[4]}"


                    if output_string not in totaldata:
                        totaldata.append(output_string)
    
    return totaldata


# Function to get the dates of the upcoming n occurrences of a given weekday
def upcoming_weekdays(day, ranges):
    # Map day names to their respective indices (0 for Monday, 1 for Tuesday, etc.)
    days_mapping = {
        'Monday': 0,
        'Tuesday': 1,
        'Wednesday': 2,
        'Thursday': 3,
        'Friday': 4,
        'Saturday': 5,
        'Sunday': 6
    }

    # Get the current date
    current_date = datetime.now().date()

    # Calculate the index of the given day
    target_day_index = days_mapping[day]

    # Find the dates of the upcoming occurrences of the given day
    upcoming_dates = []
    while len(upcoming_dates) < ranges:
        # Calculate the date of the next occurrence of the given day
        days_until_target = (target_day_index - current_date.weekday()) % 7
        next_date = current_date + timedelta(days=days_until_target)

        # Add the next occurrence date to the list
        upcoming_dates.append(next_date.strftime('%Y-%m-%d'))

        # Move to the next week
        current_date += timedelta(days=7)

    return upcoming_dates

#DELETE FROM `feedback`;
#DELETE FROM `booking id description`;
#DELETE FROM `booking list`;


async def writeData(data, db_connection, cursor):
    for val in data:
        parts = val.split(", ")
        day_of_week = upcoming_weekdays(parts[1], 12)       #CHANGE THE LAST NUMBER FOR HOW MANY WEEKS

        for day in day_of_week:
            while True:
                # Generate a new bookingRequestID
                NewbookingID = generate_random_string()

                # Checking if the ID is pre-existing
                bookingIDCheck = "SELECT `Booking ID` FROM `booking list`"
                cursor.execute(bookingIDCheck)
                currentBookingIDS = cursor.fetchall()

                # Check if the bookingRequestID is in the currentRoomIDS
                if any(NewbookingID in row for row in bookingIDCheck):
                    continue
                else:
                    break

     
            # Insert data into Booking List table
            insertDetails1 = ("""
                INSERT INTO `Booking List` (`Booking ID`, `User ID`, `Room ID`,`handler`)
                VALUES (%s, %s, %s, %s);""")
            cursor.execute(insertDetails1, (NewbookingID, parts[5], parts[0], "System"))

            #"[2R011, Monday, IFYP0023-LEC, 09:00:00, 10:00:00, Shaba]"

            # Insert the data into booking id description
            insertDetails2 = ("""
                INSERT INTO `booking id description` (`Booking ID`, `Description`, `Date`, `Start Time`, `End Time` , `Capacity`, `Comment`)
                VALUES (%s, %s, %s, %s, %s, %s, %s);""")

            cursor.execute(insertDetails2, (NewbookingID, parts[2], day, parts[3], parts[4], 30, ""))
            
            # Commit the transaction
            db_connection.commit()

    
async def datahandler(file, db_connection, cursor):
    datalist = await process_excel(file)
    await writeData(datalist, db_connection, cursor)
    return datalist
        

@app.post("/process_excel/")
async def process_excel_file(file: UploadFile = File(...), db_connection: mysql.connector.connection.MySQLConnection = Depends(get_database_connection)):
    cursor = db_connection.cursor()
    datalist = await datahandler(file, db_connection, cursor)    
    return datalist  # Await the result of process_excel

